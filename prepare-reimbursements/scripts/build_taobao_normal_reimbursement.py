#!/usr/bin/env python3
"""Build a normal reimbursement package from an edited Taobao order export."""

from __future__ import annotations

import argparse
import copy
import json
import re
import shutil
import sys
import tempfile
from dataclasses import dataclass, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DOCUMENT_TYPES = (
    "實體 Hard copy receipt/Invoice",
    "電子發票 Soft copy invoice",
    "淘寶截圖加付款紀錄 Taobao capture screen & payment record",
    "沒有 Missing",
)
DEFAULT_DOC_TYPE = DOCUMENT_TYPES[2]
DEFAULT_MISSING_REASON = "商家未提供"
DEFAULT_EVIDENCE = ["taobao_order_detail_screenshot", "payment_record_screenshot"]
EXCEL_DATE_FORMAT = "dd/mm/yyyy"
MIN_DATE_COLUMN_WIDTH = 14.0
TAOBAO_ORDER_DETAIL_URL = "https://buyertrade.taobao.com/trade/detail/trade_item_detail.htm?biz_order_id={order_no}"
ALIPAY_DETAIL_URL = "https://consumeprod.alipay.com/record/detail/simpleDetail.htm?bizType=TRADE&bizInNo={trade_no}"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


@dataclass
class OrderItem:
    name: str
    link: str
    style: str
    quantity: str
    item_amount_rmb: float | None


@dataclass
class ReimbursementOrder:
    source: str
    order_no: str
    taobao_order_detail_url: str
    alipay_trade_no: str
    alipay_detail_url: str
    date: str
    datetime: str
    shop: str
    status: str
    item_label: str
    amount_rmb: float
    shipping_rmb: float | None
    item_count: int
    items: list[OrderItem]
    document_type: str
    missing_receipt_reason: str
    evidence_required: list[str]
    claim_amount: float | None = None
    claim_currency: str = "RMB"
    payment_amount: float | None = None
    payment_currency: str = ""
    currency_review_status: str = "resolved"
    currency_note: str = ""


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_money(value: Any) -> float | None:
    text = cell_text(value)
    if not text:
        return None
    text = text.replace("￥", "").replace(",", "").replace("RMB", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def parse_datetime(value: Any) -> tuple[str, str]:
    if isinstance(value, datetime):
        return value.date().isoformat(), value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat(), value.isoformat()
    text = cell_text(value)
    if not text:
        return "", ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date().isoformat(), parsed.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return text[:10], text


def parse_excel_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = cell_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported reimbursement date: {value!r}")


def compact_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def clean_product_name(text: str) -> str:
    text = cell_text(text)
    text = re.sub(r"【优惠价】", "", text)
    text = re.sub(r"\[[^\]]{1,20}\]", "", text)
    return compact_spaces(text)


def shorten_label(text: str, max_chars: int = 24) -> str:
    text = clean_product_name(text)
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rstrip() + "..."


def make_item_label(items: list[OrderItem]) -> str:
    if not items:
        return ""
    names = [clean_product_name(item.name) for item in items if item.name]
    if not names:
        return ""
    unique_names = []
    for name in names:
        if name not in unique_names:
            unique_names.append(name)
    label = shorten_label(unique_names[0])
    if len(unique_names) > 1 or len(items) > 1:
        label += f" 等{len(items)}项"
    return label


def row_has_order_level_fields(row: list[Any]) -> bool:
    # B, C, D, J, K in the Taobao export indicate an order's first row even if
    # the user blanked out the order number to exclude it.
    return any(cell_text(row[index]) for index in (1, 2, 3, 9, 10))


def row_has_item_fields(row: list[Any]) -> bool:
    return any(cell_text(row[index]) for index in (4, 5, 6, 7, 8))


def worksheet_row_values(worksheet: Any, row_number: int) -> list[Any]:
    return [worksheet.cell(row_number, column).value for column in range(1, 12)]


def order_spans_from_merged_column_a(worksheet: Any) -> list[tuple[int, int]]:
    """Return Taobao order row spans, preferring the export's column-A merges."""
    spans: list[tuple[int, int]] = []
    covered_rows: set[int] = set()

    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_col == 1 and merged_range.max_col == 1 and merged_range.min_row >= 2:
            spans.append((merged_range.min_row, merged_range.max_row))
            covered_rows.update(range(merged_range.min_row, merged_range.max_row + 1))

    for row_number in range(2, worksheet.max_row + 1):
        if row_number in covered_rows:
            continue
        row = worksheet_row_values(worksheet, row_number)
        if cell_text(row[0]) or row_has_order_level_fields(row) or row_has_item_fields(row):
            spans.append((row_number, row_number))

    spans.sort()
    normalized: list[tuple[int, int]] = []
    last_end = 1
    for start_row, end_row in spans:
        if start_row <= last_end:
            continue
        normalized.append((start_row, end_row))
        last_end = end_row
    return normalized


def items_from_span(worksheet: Any, start_row: int, end_row: int) -> list[OrderItem]:
    items: list[OrderItem] = []
    for row_number in range(start_row, end_row + 1):
        row = worksheet_row_values(worksheet, row_number)
        if row_has_item_fields(row):
            items.append(order_item_from_row(row))
    return items


def read_taobao_orders(path: Path, include_status: str) -> tuple[list[ReimbursementOrder], dict[str, int]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active

    orders: list[ReimbursementOrder] = []
    skipped = {
        "blank_order_number": 0,
        "status": 0,
        "empty_groups": 0,
    }

    for start_row, end_row in order_spans_from_merged_column_a(worksheet):
        order_row = worksheet_row_values(worksheet, start_row)
        items = items_from_span(worksheet, start_row, end_row)
        order_no = cell_text(order_row[0])

        if not order_no:
            if row_has_order_level_fields(order_row) or items:
                skipped["blank_order_number"] += 1
            else:
                skipped["empty_groups"] += 1
            continue

        status = cell_text(order_row[2])
        if include_status and status != include_status:
            skipped["status"] += 1
            continue

        order_date, order_datetime = parse_datetime(order_row[1])
        paid = parse_money(order_row[9])
        orders.append(
            ReimbursementOrder(
                source="taobao",
                order_no=order_no,
                taobao_order_detail_url=TAOBAO_ORDER_DETAIL_URL.format(order_no=order_no),
                alipay_trade_no="",
                alipay_detail_url="",
                date=order_date,
                datetime=order_datetime,
                shop=cell_text(order_row[3]),
                status=status,
                item_label=make_item_label(items),
                amount_rmb=paid or 0.0,
                shipping_rmb=parse_money(order_row[10]),
                item_count=len(items),
                items=items,
                document_type=DEFAULT_DOC_TYPE,
                missing_receipt_reason=DEFAULT_MISSING_REASON,
                evidence_required=list(DEFAULT_EVIDENCE),
            )
        )

    return orders, skipped


def order_item_from_row(row: list[Any]) -> OrderItem:
    return OrderItem(
        name=cell_text(row[4]),
        link=cell_text(row[5]),
        style=cell_text(row[6]),
        quantity=cell_text(row[7]),
        item_amount_rmb=parse_money(row[8]),
    )


def order_to_json(order: ReimbursementOrder) -> dict[str, Any]:
    result = asdict(order)
    result["items"] = [asdict(item) for item in order.items]
    return result


def claim_amount_for_order(order: ReimbursementOrder) -> float:
    return float(order.amount_rmb if order.claim_amount is None else order.claim_amount)


def missing_receipt_reason_for_order(order: ReimbursementOrder) -> str:
    if order.document_type == DEFAULT_DOC_TYPE:
        return order.missing_receipt_reason or DEFAULT_MISSING_REASON
    if order.missing_receipt_reason == DEFAULT_MISSING_REASON:
        return ""
    return order.missing_receipt_reason


def claim_totals_by_currency(orders: list[ReimbursementOrder]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for order in orders:
        currency = (order.claim_currency or "RMB").upper()
        totals[currency] = round(totals.get(currency, 0.0) + claim_amount_for_order(order), 2)
    return totals


def currency_confirmation_reason(order: ReimbursementOrder) -> str | None:
    valid_statuses = {"resolved", "confirmed", "needs_confirmation"}
    if order.currency_review_status not in valid_statuses:
        raise ValueError(
            f"Invalid currency_review_status {order.currency_review_status!r} for "
            f"{order.source}:{order.order_no}"
        )
    if order.currency_review_status == "needs_confirmation":
        return order.currency_note or "currency confirmation required"
    if order.currency_review_status == "confirmed":
        return None

    has_payment_amount = order.payment_amount is not None
    has_payment_currency = bool(order.payment_currency)
    if has_payment_amount != has_payment_currency:
        return "payment amount and currency must be recorded together"
    if has_payment_amount:
        claim_currency = (order.claim_currency or "RMB").upper()
        payment_currency = order.payment_currency.upper()
        if claim_currency != payment_currency or abs(claim_amount_for_order(order) - float(order.payment_amount)) > 0.005:
            return "proposed claim does not match the recorded payment debit"
    return None


def validate_currency_reviews(orders: list[ReimbursementOrder]) -> None:
    pending = [(order, currency_confirmation_reason(order)) for order in orders]
    pending = [(order, reason) for order, reason in pending if reason]
    if not pending:
        return
    lines = [
        f"{order.source}:{order.order_no} purchase=RMB {order.amount_rmb:.2f} "
        f"proposed_claim={order.claim_currency} {claim_amount_for_order(order):.2f} "
        f"note={reason}"
        for order, reason in pending
    ]
    raise ValueError("Currency confirmation required before final compile:\n" + "\n".join(lines))


def write_manifest(
    path: Path,
    source_file: Path,
    orders: list[ReimbursementOrder],
    skipped: dict[str, int],
    profile: dict[str, str],
) -> None:
    payload = {
        "schema": "prepare-reimbursements.taobao-normal.v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": str(source_file),
        "profile": profile,
        "summary": {
            "included_orders": len(orders),
            "total_amount_rmb": round(sum(order.amount_rmb for order in orders), 2),
            "claim_totals": claim_totals_by_currency(orders),
            "skipped": skipped,
        },
        "orders": [order_to_json(order) for order in orders],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_review_workbook(path: Path, orders: list[ReimbursementOrder], skipped: dict[str, int]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "review"
    headers = [
        "No.",
        "Order No.",
        "Taobao Detail URL",
        "Alipay Trade No.",
        "Alipay Detail URL",
        "Date",
        "Shop",
        "Item Label",
        "Purchase Amount RMB",
        "Claim Amount",
        "Claim Currency",
        "Payment Amount",
        "Payment Currency",
        "Currency Review",
        "Currency Note",
        "Status",
        "Item Count",
        "Document Type",
        "Missing Receipt Reason",
        "Evidence Required",
        "Notes",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for index, order in enumerate(orders, 1):
        worksheet.append(
            [
                index,
                order.order_no,
                order.taobao_order_detail_url,
                order.alipay_trade_no,
                order.alipay_detail_url,
                order.date,
                order.shop,
                order.item_label,
                order.amount_rmb,
                claim_amount_for_order(order),
                order.claim_currency,
                order.payment_amount,
                order.payment_currency,
                order.currency_review_status,
                order.currency_note,
                order.status,
                order.item_count,
                order.document_type,
                order.missing_receipt_reason,
                ", ".join(order.evidence_required),
                "",
            ]
        )

    summary_row = len(orders) + 3
    worksheet.cell(summary_row, 1, "Included orders")
    worksheet.cell(summary_row, 2, len(orders))
    worksheet.cell(summary_row + 1, 1, "Purchase total RMB")
    worksheet.cell(summary_row + 1, 2, round(sum(order.amount_rmb for order in orders), 2))
    worksheet.cell(summary_row + 2, 1, "Claim totals")
    worksheet.cell(
        summary_row + 2,
        2,
        ", ".join(f"{currency} {amount:.2f}" for currency, amount in claim_totals_by_currency(orders).items()),
    )
    worksheet.cell(summary_row + 3, 1, "Skipped blank order number")
    worksheet.cell(summary_row + 3, 2, skipped.get("blank_order_number", 0))
    worksheet.cell(summary_row + 4, 1, "Skipped status")
    worksheet.cell(summary_row + 4, 2, skipped.get("status", 0))

    widths = [8, 24, 70, 30, 80, 14, 22, 36, 16, 14, 14, 16, 16, 18, 40, 12, 12, 52, 22, 46, 24]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = copy.copy(cell.alignment)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    workbook.save(path)


def find_order_export(folder: Path) -> Path:
    preferred = folder / "订单数据-报销.xlsx"
    if preferred.exists():
        return preferred
    fallback = folder / "订单数据.xlsx"
    if fallback.exists():
        return fallback
    matches = sorted(folder.glob("订单数据*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    if matches:
        return matches[0]
    raise FileNotFoundError(f"No Taobao order export found in {folder}")


def find_reimbursement_root(folder: Path) -> Path:
    current = folder.resolve()
    for parent in [current, *current.parents]:
        if parent.name == "Reimbursement":
            return parent
    return folder.parent


def is_template_candidate(path: Path, batch_folder: Path) -> bool:
    if not path.is_file():
        return False
    resolved = path.resolve()
    if resolved.is_relative_to(batch_folder):
        return False
    if any(part.lower() == "generated" for part in resolved.parts):
        return False
    if ".before-format-restore" in path.name:
        return False
    return True


def can_read_template(path: Path) -> bool:
    try:
        with path.open("rb") as handle:
            handle.read(1)
        return True
    except OSError:
        return False


def find_template(folder: Path) -> Path:
    root = find_reimbursement_root(folder)
    batch_folder = folder.resolve()
    candidates = [
        path
        for path in root.rglob("報銷清單_Reimbursement list*.xlsx")
        if is_template_candidate(path, batch_folder)
    ]
    for candidate in sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True):
        if can_read_template(candidate):
            return candidate
        print(f"warning: skipping locked template {candidate}")
    if candidates:
        raise FileNotFoundError("All previous normal reimbursement workbook templates are locked")
    else:
        raise FileNotFoundError("No previous normal reimbursement workbook template found")


def copy_row_style(worksheet: Any, source_row: int, target_row: int, max_column: int = 8) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def insert_rows_preserving_merged_cells(worksheet: Any, row: int, count: int) -> None:
    merged_ranges = list(worksheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        worksheet.unmerge_cells(str(merged_range))

    worksheet.insert_rows(row, count)

    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        if min_row >= row:
            min_row += count
            max_row += count
        worksheet.merge_cells(
            start_row=min_row,
            start_column=merged_range.min_col,
            end_row=max_row,
            end_column=merged_range.max_col,
        )


def locate_row(worksheet: Any, text: str, column: int = 1) -> int:
    for row in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row, column).value
        if isinstance(value, str) and value.strip().startswith(text):
            return row
    raise ValueError(f"Could not locate row starting with {text!r}")


def create_builtin_workbook(row_count: int) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    widths = {
        "A": 20.85546875,
        "B": 11.140625,
        "C": 31.85546875,
        "D": 13.85546875,
        "E": 14.28515625,
        "F": 25.5703125,
        "G": 58.42578125,
        "H": 32.28515625,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.merge_cells("A2:H2")
    for row in range(3, 7):
        worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)

    worksheet["A2"] = "Reimbursement list"
    worksheet["A3"] = "Name:"
    worksheet["A4"] = "Bank name:"
    worksheet["A5"] = "Bank account number:"
    worksheet["A6"] = "Leader's name"

    headers = [
        "No.",
        "Date",
        "Item",
        "Amount (HKD)",
        "Amount (RMB)",
        "Amount (Other Currencies)",
        "Document type",
        "Reason for missing receipt/invoice",
    ]
    for column, header in enumerate(headers, 1):
        worksheet.cell(7, column, header)

    total_row = 8 + max(row_count, 20)
    signature_row = total_row + 6
    worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    worksheet.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=3)
    worksheet["A" + str(total_row)] = "Total:"
    worksheet["A" + str(signature_row)] = "Leader's Signature:"
    worksheet["G" + str(signature_row)] = "Date:"

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in worksheet.iter_rows(min_row=2, max_row=signature_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border
    worksheet["A2"].font = Font(bold=True, size=14)
    worksheet["A2"].fill = title_fill
    worksheet.row_dimensions[2].height = 24
    for row in range(3, 7):
        worksheet.cell(row, 1).font = Font(bold=True)
    for cell in worksheet[7]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    worksheet.freeze_panes = "A8"
    return workbook


def load_template_or_builtin(template: Path | None, output: Path, row_count: int) -> tuple[openpyxl.Workbook, str]:
    if template:
        try:
            if template.resolve() == output.resolve():
                with tempfile.NamedTemporaryFile(suffix=template.suffix, delete=False) as handle:
                    temp_path = Path(handle.name)
                try:
                    shutil.copyfile(template, temp_path)
                    return openpyxl.load_workbook(temp_path), str(template)
                finally:
                    temp_path.unlink(missing_ok=True)
            shutil.copyfile(template, output)
            return openpyxl.load_workbook(output), str(template)
        except OSError as error:
            print(f"warning: could not read template {template}: {error}")
    return create_builtin_workbook(row_count), "built-in workbook layout"


def write_reimbursement_workbook(
    template: Path | None,
    output: Path,
    orders: list[ReimbursementOrder],
    profile: dict[str, str],
    submission_date: str,
) -> str:
    validate_currency_reviews(orders)
    invalid_document_types = sorted(
        {order.document_type for order in orders if order.document_type not in DOCUMENT_TYPES}
    )
    if invalid_document_types:
        raise ValueError(f"Unsupported document type(s): {invalid_document_types}")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook, template_used = load_template_or_builtin(template, output, len(orders))
    worksheet = workbook.active

    worksheet["B3"] = profile["name"]
    worksheet["B4"] = profile["bank"]
    worksheet["B5"] = profile["account"]
    worksheet["B6"] = profile["leader"]
    worksheet.column_dimensions["B"].width = max(
        worksheet.column_dimensions["B"].width or 0,
        MIN_DATE_COLUMN_WIDTH,
    )

    header_row = locate_row(worksheet, "No.")
    item_start = header_row + 1
    total_row = locate_row(worksheet, "Total:")
    available_rows = total_row - item_start
    needed_rows = max(0, len(orders) - available_rows)

    if needed_rows:
        insert_rows_preserving_merged_cells(worksheet, total_row, needed_rows)
        for row in range(total_row, total_row + needed_rows):
            copy_row_style(worksheet, total_row - 1, row)
        total_row += needed_rows

    for row in range(item_start, total_row):
        for column in range(1, 9):
            worksheet.cell(row, column).value = None

    for index, order in enumerate(orders, 1):
        row = item_start + index - 1
        worksheet.cell(row, 1, index)
        date_cell = worksheet.cell(row, 2)
        date_cell.value = parse_excel_date(order.date) if order.date else None
        date_cell.number_format = EXCEL_DATE_FORMAT
        worksheet.cell(row, 3, order.item_label)
        claim_amount = claim_amount_for_order(order)
        claim_currency = (order.claim_currency or "RMB").upper()
        worksheet.cell(row, 4, claim_amount if claim_currency == "HKD" else None)
        worksheet.cell(row, 5, claim_amount if claim_currency in {"RMB", "CNY"} else None)
        worksheet.cell(row, 6, claim_amount if claim_currency not in {"HKD", "RMB", "CNY"} else None)
        worksheet.cell(row, 7, order.document_type)
        worksheet.cell(row, 8, missing_receipt_reason_for_order(order))

    worksheet.data_validations.dataValidation = []
    document_type_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(DOCUMENT_TYPES) + '"',
        allow_blank=True,
    )
    worksheet.add_data_validation(document_type_validation)
    document_type_validation.add(f"G{item_start}:G{total_row - 1}")

    last_item_row = max(item_start, item_start + len(orders) - 1)
    worksheet.cell(total_row, 1, "Total:")
    worksheet.cell(total_row, 4, f"=SUM(D{item_start}:D{last_item_row})")
    worksheet.cell(total_row, 5, f"=SUM(E{item_start}:E{last_item_row})")
    worksheet.cell(total_row, 6, f"=SUM(F{item_start}:F{last_item_row})")

    signature_date = parse_excel_date(submission_date)
    for row in range(total_row + 1, min(worksheet.max_row, total_row + 12) + 1):
        if worksheet.cell(row, 7).value == "Date:":
            signature_date_cell = worksheet.cell(row, 8)
            signature_date_cell.value = signature_date
            signature_date_cell.number_format = EXCEL_DATE_FORMAT
            break

    workbook.save(output)
    return template_used


def build_profile(args: argparse.Namespace) -> dict[str, str]:
    return {
        "name": args.name,
        "bank": args.bank,
        "account": args.account,
        "leader": args.leader,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--folder", type=Path, default=Path.cwd(), help="Reimbursement batch folder")
    parser.add_argument("--orders", type=Path, help="Edited Taobao order export. Defaults to 订单数据-报销.xlsx in folder")
    parser.add_argument("--out-dir", type=Path, help="Output directory. Defaults to <folder>/generated")
    parser.add_argument("--template", type=Path, help="Previous normal reimbursement workbook to use as template")
    parser.add_argument("--submission-date", default=date.today().isoformat(), help="YYYY-MM-DD date for output filename and signature")
    parser.add_argument("--include-status", default="交易成功", help="Only include orders with this status; blank includes all statuses")
    parser.add_argument("--name", default="Li Gaoyang")
    parser.add_argument("--bank", default="HSBC")
    parser.add_argument("--account", default="592-251326-833")
    parser.add_argument("--leader", default="Chen Wei")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    folder = args.folder.resolve()
    orders_path = (args.orders or find_order_export(folder)).resolve()
    out_dir = (args.out_dir or folder / "generated").resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    try:
        template = (args.template or find_template(folder)).resolve()
    except FileNotFoundError:
        template = None
    profile = build_profile(args)

    orders, skipped = read_taobao_orders(orders_path, args.include_status)

    manifest_path = out_dir / "reimbursement-manifest.json"
    review_path = out_dir / "reimbursement-review.xlsx"
    workbook_path = out_dir / f"報銷清單_Reimbursement list {args.name} {args.submission_date}.xlsx"

    write_manifest(manifest_path, orders_path, orders, skipped, profile)
    write_review_workbook(review_path, orders, skipped)
    template_used = write_reimbursement_workbook(template, workbook_path, orders, profile, args.submission_date)

    print(f"source: {orders_path}")
    print(f"template: {template_used}")
    print(f"included orders: {len(orders)}")
    print(f"total amount RMB: {sum(order.amount_rmb for order in orders):.2f}")
    print(f"skipped blank order number: {skipped.get('blank_order_number', 0)}")
    print(f"skipped status: {skipped.get('status', 0)}")
    print(f"manifest: {manifest_path}")
    print(f"review workbook: {review_path}")
    print(f"reimbursement workbook: {workbook_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
