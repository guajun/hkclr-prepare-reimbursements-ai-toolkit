#!/usr/bin/env python3
"""Compile travel reimbursement outputs from SQLite state."""

from __future__ import annotations

import argparse
import copy
import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from build_taobao_normal_reimbursement import (
    copy_row_style,
    insert_rows_preserving_merged_cells,
    locate_row,
)
from prepare_reimbursements import state_db
from sync_travel_reimbursement_state import (
    DEFAULT_DB_NAME,
    EXPENSE_COLUMNS,
    find_travel_workbook,
    load_workbook_from_onedrive_safe,
    relative_to_folder,
)

COMPILE_SCHEMA = "prepare-reimbursements.travel.compile-from-state.v1"
EVIDENCE_SUMMARY_SCHEMA = "prepare-reimbursements.travel-evidence-summary.v1"
EXCEL_DATE_FORMAT = "dd/mm/yyyy"
MIN_DATE_COLUMN_WIDTH = 14.0

CATEGORY_COLUMN = {
    (category, currency): column
    for column, category, _category_label, currency in EXPENSE_COLUMNS
}


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def load_json_field(value: str | None, default: Any) -> Any:
    return state_db.json_loads(value, default)


def load_batch(connection: Any, batch_folder: Path) -> Any:
    row = connection.execute(
        "SELECT * FROM batches WHERE batch_folder = ?",
        (str(batch_folder),),
    ).fetchone()
    if row is None:
        raise FileNotFoundError(f"No SQLite batch state found for {batch_folder}")
    return row


def load_expense_rows(connection: Any, batch_id: int) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT * FROM travel_expense_rows
        WHERE batch_id = ?
        ORDER BY source_row_index, category, currency
        """,
        (batch_id,),
    ).fetchall()
    return [
        {
            "source_row_index": int(row["source_row_index"]),
            "expense_date": row["expense_date"] or "",
            "destination": row["destination"] or "",
            "category": row["category"],
            "category_label": row["category_label"],
            "currency": row["currency"],
            "amount": float(row["amount"] or 0),
            "raw": load_json_field(row["raw_json"], {}),
        }
        for row in rows
    ]


def load_itinerary_rows(connection: Any, batch_id: int) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT * FROM travel_itinerary_rows
        WHERE batch_id = ?
        ORDER BY itinerary_index
        """,
        (batch_id,),
    ).fetchall()
    return [
        {
            "itinerary_index": int(row["itinerary_index"]),
            "trip_date": row["trip_date"] or "",
            "origin": row["origin"] or "",
            "destination": row["destination"] or "",
            "purpose": row["purpose"] or "",
            "raw": load_json_field(row["raw_json"], {}),
        }
        for row in rows
    ]


def load_travel_evidence(connection: Any, batch_id: int) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT * FROM travel_evidence_files
        WHERE batch_id = ?
        ORDER BY evidence_index, evidence_kind
        """,
        (batch_id,),
    ).fetchall()
    return [
        {
            "index": int(row["evidence_index"]),
            "kind": row["evidence_kind"],
            "relative_path": row["relative_path"],
            "path": row["actual_path"],
            "file_name": row["file_name"],
            "file_size": row["file_size"],
            "sha256": row["sha256"],
            "size": [row["width"], row["height"]] if row["width"] and row["height"] else None,
            "status": row["validation_status"],
            "warnings": load_json_field(row["warnings_json"], []),
            "details": load_json_field(row["details_json"], {}),
        }
        for row in rows
    ]


def parse_excel_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported travel reimbursement date: {value!r}")


def group_expenses(expenses: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[int, dict[str, Any]] = {}
    for expense in expenses:
        source_row = int(expense["source_row_index"])
        row = grouped.setdefault(
            source_row,
            {
                "source_row_index": source_row,
                "expense_date": expense["expense_date"],
                "destination": expense["destination"],
                "amounts": {},
            },
        )
        row["amounts"][(expense["category"], expense["currency"])] = expense["amount"]
    return [grouped[key] for key in sorted(grouped)]


def copy_workbook_template(template: Path, output: Path) -> Any:
    output.parent.mkdir(parents=True, exist_ok=True)
    if template.resolve() == output.resolve():
        return load_workbook_from_onedrive_safe(template)
    shutil.copyfile(template, output)
    return openpyxl.load_workbook(output)


def create_builtin_travel_workbook(expense_rows: int, itinerary_rows: int) -> Any:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "差旅報銷清單"
    sheet.merge_cells("A1:R2")
    sheet.merge_cells("A3:H5")
    sheet.merge_cells("I3:R4")
    sheet.merge_cells("I5:R5")
    sheet.merge_cells("A6:R6")
    sheet["A1"] = "Reimbursement for Travel Expenses  差旅費報銷清單"
    sheet["A3"] = "Applicant:"
    sheet["I3"] = "Bank:"
    sheet["I5"] = "Bank Account Number:"
    sheet["A6"] = "Purpose of the trip:"
    sheet["A7"] = "Date"
    sheet["B7"] = "Destination"
    for column, _category, category_label, currency in EXPENSE_COLUMNS:
        sheet.cell(8, column, category_label)
        sheet.cell(9, column, currency)
    total_row = 10 + max(expense_rows, 1)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    sheet.cell(total_row, 1, "Total:")

    itinerary = workbook.create_sheet("行程資料列表")
    itinerary.append(["", "日期", "出發地", "目的地", "原因/目的"])
    for _ in range(max(itinerary_rows, 1)):
        itinerary.append(["", "", "", "", ""])

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = border
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    return workbook


def load_template_or_builtin(template: Path | None, output: Path, expense_rows: int, itinerary_rows: int) -> tuple[Any, str]:
    if template and template.exists():
        try:
            return copy_workbook_template(template, output), str(template)
        except OSError as error:
            print(f"warning: could not read travel template {template}: {error}")
    workbook = create_builtin_travel_workbook(expense_rows, itinerary_rows)
    return workbook, "built-in travel workbook layout"


def ensure_expense_capacity(worksheet: Any, row_count: int) -> tuple[int, int]:
    data_start = 10
    total_row = locate_row(worksheet, "Total:")
    available_rows = total_row - data_start
    needed_rows = max(0, row_count - available_rows)
    if needed_rows:
        insert_rows_preserving_merged_cells(worksheet, total_row, needed_rows)
        for row in range(total_row, total_row + needed_rows):
            copy_row_style(worksheet, total_row - 1, row, max_column=18)
        total_row += needed_rows
    return data_start, total_row


def write_expense_sheet(worksheet: Any, profile: dict[str, Any], expenses: list[dict[str, Any]]) -> None:
    rows = group_expenses(expenses)
    worksheet.column_dimensions["A"].width = max(
        worksheet.column_dimensions["A"].width or 0,
        MIN_DATE_COLUMN_WIDTH,
    )
    worksheet["A3"] = f"Applicant: {profile.get('name', '')}"
    worksheet["I3"] = f"Bank: {profile.get('bank', '')}"
    worksheet["I5"] = f"Bank Account Number: {profile.get('account', '')}"
    worksheet["A6"] = f"Purpose of the trip: {profile.get('purpose', '')}"

    data_start, total_row = ensure_expense_capacity(worksheet, len(rows))
    for row_number in range(data_start, total_row):
        for column in range(1, 18):
            worksheet.cell(row_number, column).value = None

    for offset, source_row in enumerate(rows):
        row_number = data_start + offset
        date_cell = worksheet.cell(row_number, 1)
        date_cell.value = parse_excel_date(source_row["expense_date"])
        date_cell.number_format = EXCEL_DATE_FORMAT
        worksheet.cell(row_number, 2, source_row["destination"])
        for key, amount in source_row["amounts"].items():
            column = CATEGORY_COLUMN.get(key)
            if column:
                worksheet.cell(row_number, column, amount)

    last_data_row = max(data_start, data_start + len(rows) - 1)
    worksheet.cell(total_row, 1, "Total:")
    for column in range(3, 18):
        letter = get_column_letter(column)
        worksheet.cell(total_row, column, f"=SUM({letter}{data_start}:{letter}{last_data_row})")


def copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def ensure_itinerary_capacity(worksheet: Any, row_count: int) -> None:
    needed_max_row = row_count + 1
    if worksheet.max_row >= needed_max_row:
        return
    template_row = max(2, worksheet.max_row)
    for row_number in range(worksheet.max_row + 1, needed_max_row + 1):
        worksheet.append(["", "", "", "", ""])
        worksheet.row_dimensions[row_number].height = worksheet.row_dimensions[template_row].height
        for column in range(1, 6):
            copy_cell_style(worksheet.cell(template_row, column), worksheet.cell(row_number, column))


def write_itinerary_sheet(worksheet: Any, itinerary: list[dict[str, Any]]) -> None:
    ensure_itinerary_capacity(worksheet, len(itinerary))
    worksheet.column_dimensions["B"].width = max(
        worksheet.column_dimensions["B"].width or 0,
        MIN_DATE_COLUMN_WIDTH,
    )
    for row_number in range(2, worksheet.max_row + 1):
        for column in range(1, 6):
            worksheet.cell(row_number, column).value = None

    for offset, row in enumerate(itinerary, 2):
        worksheet.cell(offset, 1, row["itinerary_index"])
        date_cell = worksheet.cell(offset, 2)
        date_cell.value = parse_excel_date(row["trip_date"])
        date_cell.number_format = EXCEL_DATE_FORMAT
        worksheet.cell(offset, 3, row["origin"])
        worksheet.cell(offset, 4, row["destination"])
        worksheet.cell(offset, 5, row["purpose"])


def artifact_for(path: Path, batch_folder: Path, kind: str, details: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "artifact_kind": kind,
        "path": str(path),
        "relative_path": relative_to_folder(path, batch_folder),
        "file_size": path.stat().st_size if path.exists() and path.is_file() else None,
        "sha256": state_db.sha256_file(path) if path.exists() and path.is_file() else None,
        "details": details or {},
    }


def write_evidence_summary(path: Path, evidence: list[dict[str, Any]]) -> dict[str, Any]:
    summary = {
        "schema": EVIDENCE_SUMMARY_SCHEMA,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "evidence_records": len(evidence),
        "valid_evidence_records": sum(1 for row in evidence if row["status"] == "valid"),
        "warning_records": sum(1 for row in evidence if row["warnings"]),
        "evidence": evidence,
    }
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def compile_outputs(
    *,
    batch_folder: Path,
    db_path: Path,
    out_dir: Path,
    workbook_output: Path | None,
    template: Path | None,
    submission_date: str,
    overrides: dict[str, str],
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    with state_db.connect(db_path) as connection:
        batch = load_batch(connection, batch_folder)
        batch_id = int(batch["id"])
        profile = load_json_field(batch["profile_json"], {})
        profile.update({key: value for key, value in overrides.items() if value})
        expenses = load_expense_rows(connection, batch_id)
        itinerary = load_itinerary_rows(connection, batch_id)
        evidence = load_travel_evidence(connection, batch_id)

        template_path = template
        batch_summary = load_json_field(batch["summary_json"], {})
        travel_summary = batch_summary.get("travel") or {}
        if template_path is None and travel_summary.get("source_workbook"):
            template_path = Path(travel_summary["source_workbook"])
        if template_path is None and batch["reimbursement_type"] == "travel" and batch["source_export_path"]:
            template_path = Path(batch["source_export_path"])
        if template_path is None:
            template_path = find_travel_workbook(batch_folder)
        output_name = f"差旅報銷清單_行程資料列表Reimbursement for travel expenses - {profile.get('name', 'Li Gaoyang')} {submission_date}.xlsx"
        workbook_path = workbook_output or batch_folder / output_name
        evidence_summary_path = out_dir / "travel-evidence-summary.json"
        compile_summary_path = out_dir / "travel-reimbursement-state-compile-summary.json"

        workbook, template_used = load_template_or_builtin(
            template_path,
            workbook_path,
            expense_rows=len(group_expenses(expenses)),
            itinerary_rows=len(itinerary),
        )
        expense_sheet = workbook["差旅報銷清單"]
        itinerary_sheet = workbook["行程資料列表"]
        write_expense_sheet(expense_sheet, profile, expenses)
        write_itinerary_sheet(itinerary_sheet, itinerary)
        workbook.save(workbook_path)

        evidence_summary = write_evidence_summary(evidence_summary_path, evidence)
        compile_summary = {
            "schema": COMPILE_SCHEMA,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "database": str(db_path),
            "batch": str(batch_folder),
            "template": template_used,
            "travel_workbook": str(workbook_path),
            "submission_date": submission_date,
            "expense_records": len(expenses),
            "expense_source_rows": len(group_expenses(expenses)),
            "itinerary_rows": len(itinerary),
            "evidence_summary": str(evidence_summary_path),
            "valid_evidence_records": evidence_summary["valid_evidence_records"],
        }
        compile_summary_path.write_text(json.dumps(compile_summary, ensure_ascii=False, indent=2), encoding="utf-8")

        state_db.upsert_artifacts(
            connection,
            batch_id=batch_id,
            artifacts=[
                artifact_for(workbook_path, batch_folder, "travel_reimbursement_workbook"),
                artifact_for(evidence_summary_path, batch_folder, "travel_evidence_summary"),
                artifact_for(compile_summary_path, batch_folder, "travel_state_compile_summary"),
            ],
        )
        connection.commit()
    return compile_summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--folder", type=Path, required=True, help="Reimbursement batch folder")
    parser.add_argument("--db", type=Path, help="SQLite state database. Defaults to <folder>/generated/reimbursement-state.sqlite3")
    parser.add_argument("--out-dir", type=Path, help="Output directory. Defaults to <folder>/generated")
    parser.add_argument(
        "--workbook-output",
        type=Path,
        help="Final travel workbook path. Defaults to the batch folder beside the normal reimbursement workbook",
    )
    parser.add_argument("--template", type=Path, help="Travel reimbursement workbook to use as template")
    parser.add_argument("--submission-date", default=date.today().isoformat(), help="YYYY-MM-DD date for output filename")
    parser.add_argument("--name")
    parser.add_argument("--bank")
    parser.add_argument("--account")
    parser.add_argument("--purpose")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    batch_folder = args.folder.resolve()
    out_dir = (args.out_dir or batch_folder / "generated").resolve()
    db_path = (args.db or batch_folder / "generated" / DEFAULT_DB_NAME).resolve()
    overrides = {
        "name": args.name,
        "bank": args.bank,
        "account": args.account,
        "purpose": args.purpose,
    }
    result = compile_outputs(
        batch_folder=batch_folder,
        db_path=db_path,
        out_dir=out_dir,
        workbook_output=args.workbook_output.resolve() if args.workbook_output else None,
        template=args.template.resolve() if args.template else None,
        submission_date=args.submission_date,
        overrides=overrides,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
