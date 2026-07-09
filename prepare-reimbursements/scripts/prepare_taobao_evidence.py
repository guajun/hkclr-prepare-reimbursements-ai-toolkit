#!/usr/bin/env python3
"""Prepare Taobao evidence folders and capture checklists from a manifest."""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from PIL import Image, ImageChops, ImageStat, UnidentifiedImageError
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ALIPAY_DETAIL_URL = "https://consumeprod.alipay.com/record/detail/simpleDetail.htm?bizType=TRADE&bizInNo={trade_no}"
MIN_ORDER_SCREENSHOT_WIDTH = 900
MIN_ORDER_SCREENSHOT_HEIGHT = 500
MIN_PAYMENT_SCREENSHOT_WIDTH = 800
APPROVED_PAYMENT_SCREENSHOT_SIZES = {(820, 777), (911, 777), (1425, 801), (1521, 633), (1521, 688), (1536, 639)}
PRINT_FLAT_ROOT = Path("generated") / "print-flat" / "taobao"
KNOWN_BAD_TILED_RANGES = [
    {
        "width": (4200, 4350),
        "height": (2350, 2450),
        "warning": "known bad Codex in-app browser 2x2 tiled capture around 4276x2404",
    },
]


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def existing_any(folder: Path, patterns: list[str]) -> list[Path]:
    hits: list[Path] = []
    for pattern in patterns:
        hits.extend(folder.glob(pattern))
    return sorted(set(hits), key=lambda path: path.name)


def image_size(path: Path) -> tuple[int, int] | None:
    try:
        with path.open("rb") as handle:
            header = handle.read(24)
            if header.startswith(b"\x89PNG\r\n\x1a\n") and len(header) >= 24:
                width = int.from_bytes(header[16:20], "big")
                height = int.from_bytes(header[20:24], "big")
                return width, height

            if header[:2] != b"\xff\xd8":
                return None

            handle.seek(2)
            while True:
                marker_start = handle.read(1)
                while marker_start and marker_start != b"\xff":
                    marker_start = handle.read(1)
                marker = handle.read(1)
                while marker == b"\xff":
                    marker = handle.read(1)
                if not marker:
                    return None
                marker_value = marker[0]
                if marker_value in {0xD8, 0xD9}:
                    continue
                length_bytes = handle.read(2)
                if len(length_bytes) != 2:
                    return None
                length = int.from_bytes(length_bytes, "big")
                if length < 2:
                    return None
                if marker_value in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
                    frame = handle.read(5)
                    if len(frame) != 5:
                        return None
                    height = int.from_bytes(frame[1:3], "big")
                    width = int.from_bytes(frame[3:5], "big")
                    return width, height
                handle.seek(length - 2, 1)
    except OSError:
        return None


def known_bad_tiled_warning(width: int, height: int) -> str:
    for item in KNOWN_BAD_TILED_RANGES:
        width_min, width_max = item["width"]
        height_min, height_max = item["height"]
        if width_min <= width <= width_max and height_min <= height <= height_max:
            return str(item["warning"])
    return ""


def mean_difference(first: Image.Image, second: Image.Image) -> float:
    sample_size = (160, 160)
    left = first.convert("L").resize(sample_size)
    right = second.convert("L").resize(sample_size)
    diff = ImageChops.difference(left, right)
    return float(ImageStat.Stat(diff).mean[0])


def tiled_capture_warnings(path: Path) -> list[str]:
    warnings: list[str] = []
    try:
        with Image.open(path) as image:
            width, height = image.size
            bad_warning = known_bad_tiled_warning(width, height)
            if bad_warning:
                warnings.append(f"{path.name}: {bad_warning}")

            if width >= 1200 and height >= 1200:
                half_width = width // 2
                half_height = height // 2
                top_left = image.crop((0, 0, half_width, half_height))
                top_right = image.crop((half_width, 0, half_width * 2, half_height))
                bottom_left = image.crop((0, half_height, half_width, half_height * 2))
                if mean_difference(top_left, top_right) < 2.0 and mean_difference(top_left, bottom_left) < 2.0:
                    warnings.append(f"{path.name}: looks like a duplicated 2x2 tiled browser screenshot")

            if width >= 900 and height >= 1800:
                band_height = min(700, height // 3)
                first_band = image.crop((0, 0, width, band_height))
                second_band = image.crop((0, band_height, width, band_height * 2))
                if mean_difference(first_band, second_band) < 2.0:
                    warnings.append(f"{path.name}: adjacent vertical bands look duplicated")
    except (OSError, UnidentifiedImageError) as error:
        warnings.append(f"{path.name}: could not inspect image quality: {error}")
    return warnings


def order_screenshot_warnings(paths: list[Path]) -> list[str]:
    warnings: list[str] = []
    for path in paths:
        size = image_size(path)
        if not size:
            warnings.append(f"{path.name}: could not read image dimensions")
            continue
        width, height = size
        if width < MIN_ORDER_SCREENSHOT_WIDTH or height < MIN_ORDER_SCREENSHOT_HEIGHT:
            warnings.append(
                f"{path.name}: size {width}x{height}px is too small for a reliable Taobao order-detail screenshot"
            )
        warnings.extend(tiled_capture_warnings(path))
    return warnings


def payment_screenshot_warnings(paths: list[Path]) -> list[str]:
    warnings: list[str] = []
    for path in paths:
        size = image_size(path)
        if not size:
            continue
        width, height = size
        if width < MIN_PAYMENT_SCREENSHOT_WIDTH:
            warnings.append(
                f"{path.name}: width {width}px; verify the right-side '= 实付金额' amount and payment method are not cropped"
            )
        if (width, height) not in APPROVED_PAYMENT_SCREENSHOT_SIZES:
            approved = ", ".join(f"{approved_width}x{approved_height}" for approved_width, approved_height in sorted(APPROVED_PAYMENT_SCREENSHOT_SIZES))
            warnings.append(
                f"{path.name}: size {width}x{height}px is not an approved Alipay screenshot preset ({approved}); rerun normalization or inspect manually"
            )
    return warnings


def make_order_folder(evidence_root: Path, index: int, order_no: str) -> Path:
    folder = evidence_root / f"{index:02d}_{order_no}"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def alipay_detail_url(order: dict[str, Any]) -> str:
    if order.get("alipay_detail_url"):
        return str(order["alipay_detail_url"])
    trade_no = str(order.get("alipay_trade_no") or "").strip()
    if trade_no:
        return ALIPAY_DETAIL_URL.format(trade_no=trade_no)
    return ""


def write_folder_note(folder: Path, order: dict[str, Any], index: int) -> tuple[str, str, str]:
    order_no = order["order_no"]
    order_file = f"{index:02d}_{order_no}_taobao_order_detail.png"
    payment_file = f"{index:02d}_{order_no}_payment_record.png"
    combined_file = f"{index:02d}_{order_no}_combined_receipt.png"
    taobao_url = order.get("taobao_order_detail_url") or ""
    alipay_trade_no = order.get("alipay_trade_no") or ""
    alipay_url = alipay_detail_url(order)
    note_path = folder / "_放截图到这里.txt"
    note_path.write_text(
        "\n".join(
            [
                f"订单号: {order_no}",
                f"日期: {order['date']}",
                f"店铺: {order['shop']}",
                f"金额: RMB {order['amount_rmb']}",
                f"物品: {order['item_label']}",
                f"淘宝详情页: {taobao_url}",
                f"支付宝交易号: {alipay_trade_no or '先从淘宝详情页提取'}",
                f"支付宝详情页: {alipay_url or '提取支付宝交易号后生成'}",
                "",
                "需要放入:",
                f"1. 淘宝订单详情截图: {order_file}",
                "2. 从淘宝详情页提取字段: 支付宝交易号",
                f"3. 打开支付宝详情页并截图: {payment_file}",
                "   验收: 付款图必须显示交易成功、流水号、时间、订单金额、= 实付金额、实付金额数字和付款方式。",
                "   如浏览器截图出现重复平铺，先保存到 _raw_payment_screenshots，再运行 normalize_alipay_payment_screenshots.py；不要临场手裁。",
                f"4. 可选合成凭证: {combined_file}",
            ]
        ),
        encoding="utf-8",
    )
    return order_file, payment_file, combined_file


def build_records(orders: list[dict[str, Any]], evidence_root: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for index, order in enumerate(orders, 1):
        folder = make_order_folder(evidence_root, index, order["order_no"])
        order_file, payment_file, combined_file = write_folder_note(folder, order, index)
        order_hits = [
            path
            for path in existing_any(folder, ["*order*.*", "*taobao*.*", "*淘宝*.*", "*订单*.*"])
            if path.name != "_放截图到这里.txt"
        ]
        payment_hits = [
            path
            for path in existing_any(folder, ["*payment*.*", "*alipay*.*", "*支付宝*.*", "*付款*.*"])
            if path.name != "_放截图到这里.txt"
        ]
        combined_hits = [
            path
            for path in existing_any(folder, ["*combined*.*", "*receipt*.*", "*凭证*.*", "*合成*.*"])
            if path.name != "_放截图到这里.txt"
        ]
        records.append(
            {
                "index": index,
                "folder": folder,
                "order": order,
                "order_file": order_file,
                "payment_file": payment_file,
                "combined_file": combined_file,
                "order_hits": [path.name for path in order_hits],
                "payment_hits": [path.name for path in payment_hits],
                "order_image_warnings": order_screenshot_warnings(order_hits),
                "payment_image_warnings": payment_screenshot_warnings(payment_hits),
                "combined_hits": [path.name for path in combined_hits],
            }
        )
    return records


def style_worksheet(worksheet: Any, widths: list[int], header_color: str = "4F81BD") -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"


def write_checklist(path: Path, records: list[dict[str, Any]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "evidence-checklist"
    headers = [
        "No.",
        "Order No.",
        "Taobao Detail URL",
        "Alipay Trade No.",
        "Alipay Detail URL",
        "Date",
        "Shop",
        "Item Label",
        "Amount RMB",
        "Item Count",
        "Folder",
        "Taobao Order Screenshot Filename",
        "Payment Screenshot Filename",
        "Order Screenshot Found",
        "Order Screenshot Check",
        "Payment Screenshot Found",
        "Payment Screenshot Check",
        "Combined Receipt Found",
        "First Item Link",
    ]
    worksheet.append(headers)
    for record in records:
        order = record["order"]
        first_link = ""
        if order.get("items"):
            first_link = order["items"][0].get("link") or ""
        worksheet.append(
            [
                record["index"],
                order["order_no"],
                order.get("taobao_order_detail_url") or "",
                order.get("alipay_trade_no") or "",
                alipay_detail_url(order),
                order["date"],
                order["shop"],
                order["item_label"],
                order["amount_rmb"],
                order["item_count"],
                str(record["folder"]),
                record["order_file"],
                record["payment_file"],
                "YES" if record["order_hits"] and not record["order_image_warnings"] else "NO",
                "; ".join(record["order_image_warnings"]),
                "YES" if record["payment_hits"] and not record["payment_image_warnings"] else "NO",
                "; ".join(record["payment_image_warnings"]),
                "YES" if record["combined_hits"] else "NO",
                first_link,
            ]
        )
        if first_link:
            worksheet.cell(worksheet.max_row, 19).hyperlink = first_link
            worksheet.cell(worksheet.max_row, 19).style = "Hyperlink"
        taobao_url = order.get("taobao_order_detail_url") or ""
        if taobao_url:
            worksheet.cell(worksheet.max_row, 3).hyperlink = taobao_url
            worksheet.cell(worksheet.max_row, 3).style = "Hyperlink"
        payment_url = alipay_detail_url(order)
        if payment_url:
            worksheet.cell(worksheet.max_row, 5).hyperlink = payment_url
            worksheet.cell(worksheet.max_row, 5).style = "Hyperlink"
    style_worksheet(worksheet, [6, 24, 70, 30, 80, 12, 22, 40, 12, 10, 72, 42, 42, 18, 34, 18, 34, 18, 50])

    items = workbook.create_sheet("items")
    items.append(["Order No.", "Order Index", "Item Name", "Style", "Quantity", "Item Amount RMB", "Product Link"])
    for record in records:
        order = record["order"]
        for item in order.get("items", []):
            items.append(
                [
                    order["order_no"],
                    record["index"],
                    item.get("name"),
                    item.get("style"),
                    item.get("quantity"),
                    item.get("item_amount_rmb"),
                    item.get("link"),
                ]
            )
            link = item.get("link") or ""
            if link:
                items.cell(items.max_row, 7).hyperlink = link
                items.cell(items.max_row, 7).style = "Hyperlink"
    style_worksheet(items, [24, 10, 64, 30, 10, 16, 60], "70AD47")
    workbook.save(path)


def write_capture_queue(path: Path, batch_folder: Path, records: list[dict[str, Any]]) -> None:
    lines = [
        "# Taobao Evidence Capture Queue",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Batch: `{batch_folder}`",
        "",
        "For each order, capture:",
        "",
        "1. Taobao order detail screenshot showing order number, items, shop, date, and paid amount.",
        "2. Extract the Alipay trade number from the Taobao order detail page. Prefer the field labelled `支付宝交易号`.",
        "3. Open `https://consumeprod.alipay.com/record/detail/simpleDetail.htm?bizType=TRADE&bizInNo=<支付宝交易号>` directly and capture the payment record screenshot.",
        "4. Save the raw browser screenshot under `_raw_payment_screenshots`, then run `scripts/normalize_alipay_payment_screenshots.py`. Accept only approved preset outputs and inspect the contact sheet before continuing.",
        "5. Optional combined receipt image after pasting the narrow payment record into the order screenshot.",
        "",
    ]
    for record in records:
        order = record["order"]
        lines.extend(
            [
                f"## {record['index']:02d}. {order['order_no']}",
                "",
                f"- Date: {order['date']}",
                f"- Shop: {order['shop']}",
                f"- Amount: RMB {order['amount_rmb']}",
                f"- Item label: {order['item_label']}",
                f"- Folder: `{record['folder']}`",
                f"- Taobao detail URL: {order.get('taobao_order_detail_url') or ''}",
                f"- Alipay trade no: {order.get('alipay_trade_no') or '(extract from Taobao detail page)'}",
                f"- Alipay detail URL: {alipay_detail_url(order) or '(generated after extraction)'}",
                f"- Save Taobao screenshot as: `{record['order_file']}`",
                f"- Save payment screenshot as: `{record['payment_file']}`",
                f"- Optional combined image: `{record['combined_file']}`",
            ]
        )
        if order.get("items"):
            lines.append("- Product links:")
            for item in order["items"][:5]:
                lines.append(f"  - {item.get('name') or ''} / {item.get('style') or ''}: {item.get('link') or ''}")
            if len(order["items"]) > 5:
                lines.append(f"  - ... {len(order['items']) - 5} more items in the checklist workbook")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def reset_flat_folder(path: Path) -> list[str]:
    warnings: list[str] = []
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_file() or child.is_symlink():
            child.unlink()
        else:
            warnings.append(f"Skipped non-file item in print-flat folder: {child}")
    return warnings


def link_print_file(source: Path, destination: Path) -> tuple[str, str]:
    relative_source = os.path.relpath(source, destination.parent)
    try:
        destination.symlink_to(relative_source)
        return "symlink", ""
    except OSError as symlink_error:
        try:
            os.link(source, destination)
            return "hardlink", f"Symlink failed for {destination.name}; created hardlink instead: {symlink_error}"
        except OSError as hardlink_error:
            return "missing", f"Could not create print-flat link for {source}: symlink error: {symlink_error}; hardlink error: {hardlink_error}"


def write_print_flat_folder(path: Path, records: list[dict[str, Any]]) -> dict[str, Any]:
    warnings = reset_flat_folder(path)
    links: list[dict[str, Any]] = []
    sequence = 1
    for record in records:
        order = record["order"]
        printable_files = [
            ("taobao_order_detail", record["folder"] / record["order_file"]),
            ("payment_record", record["folder"] / record["payment_file"]),
        ]
        for label, source in printable_files:
            if not source.exists():
                warnings.append(f"Missing printable source for {record['index']:02d} {order['order_no']}: {source.name}")
                continue
            destination = path / f"{sequence:03d}_{record['index']:02d}_{order['order_no']}_{label}{source.suffix.lower()}"
            link_type, warning = link_print_file(source, destination)
            if warning:
                warnings.append(warning)
            if link_type != "missing":
                links.append(
                    {
                        "sequence": sequence,
                        "index": record["index"],
                        "order_no": order["order_no"],
                        "kind": label,
                        "link_type": link_type,
                        "path": str(destination),
                        "target": str(source),
                    }
                )
                sequence += 1
    return {"folder": str(path), "links": links, "warnings": warnings}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--folder", type=Path, required=True, help="Reimbursement batch folder")
    parser.add_argument("--manifest", type=Path, help="Manifest path. Defaults to <folder>/generated/reimbursement-manifest.json")
    parser.add_argument("--evidence-root", type=Path, help="Evidence root. Defaults to <folder>/物品/taobao")
    parser.add_argument("--no-print-flat", action="store_true", help="Do not refresh the generated/print-flat/taobao folder for bulk printing")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    batch_folder = args.folder.resolve()
    generated = batch_folder / "generated"
    manifest_path = (args.manifest or generated / "reimbursement-manifest.json").resolve()
    evidence_root = (args.evidence_root or batch_folder / "物品" / "taobao").resolve()
    evidence_root.mkdir(parents=True, exist_ok=True)

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    records = build_records(manifest["orders"], evidence_root)

    checklist_path = generated / "taobao-evidence-checklist.xlsx"
    queue_path = generated / "taobao-evidence-capture-queue.md"
    summary_path = generated / "taobao-evidence-summary.json"

    write_checklist(checklist_path, records)
    write_capture_queue(queue_path, batch_folder, records)
    print_flat = None
    if not args.no_print_flat:
        print_flat = write_print_flat_folder(batch_folder / PRINT_FLAT_ROOT, records)

    summary = {
        "evidence_root": str(evidence_root),
        "orders": len(records),
        "order_screenshots_present": sum(1 for record in records if record["order_hits"]),
        "order_screenshots_found": sum(1 for record in records if record["order_hits"] and not record["order_image_warnings"]),
        "order_screenshot_warnings": [
            {"folder": str(record["folder"]), "warnings": record["order_image_warnings"]}
            for record in records
            if record["order_image_warnings"]
        ],
        "payment_screenshots_present": sum(1 for record in records if record["payment_hits"]),
        "payment_screenshots_found": sum(1 for record in records if record["payment_hits"] and not record["payment_image_warnings"]),
        "payment_screenshot_warnings": [
            {"folder": str(record["folder"]), "warnings": record["payment_image_warnings"]}
            for record in records
            if record["payment_image_warnings"]
        ],
        "combined_receipts_found": sum(1 for record in records if record["combined_hits"]),
        "checklist": str(checklist_path),
        "queue": str(queue_path),
    }
    if print_flat:
        summary["print_flat_folder"] = print_flat["folder"]
        summary["print_flat_links"] = len(print_flat["links"])
        summary["print_flat_warnings"] = print_flat["warnings"]
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
