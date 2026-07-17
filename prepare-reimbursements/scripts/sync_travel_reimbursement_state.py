#!/usr/bin/env python3
"""Sync a travel reimbursement workbook and evidence files into SQLite state."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from prepare_reimbursements import state_db

DEFAULT_DB_NAME = "reimbursement-state.sqlite3"
DEFAULT_SNAPSHOT_NAME = "reimbursement-state.snapshot.json"
MANIFEST_NAME = "travel-reimbursement-manifest.json"

EXPENSE_COLUMNS = [
    (3, "transport", "Flight/Vessel/Train/Car", "HKD"),
    (4, "transport", "Flight/Vessel/Train/Car", "RMB"),
    (5, "transport", "Flight/Vessel/Train/Car", "Other"),
    (6, "hotel", "Hotel", "HKD"),
    (7, "hotel", "Hotel", "RMB"),
    (8, "hotel", "Hotel", "Other"),
    (9, "conference_fee", "Conference Fee", "HKD"),
    (10, "conference_fee", "Conference Fee", "RMB"),
    (11, "conference_fee", "Conference Fee", "Other"),
    (12, "meal", "Meal", "HKD"),
    (13, "meal", "Meal", "RMB"),
    (14, "meal", "Meal", "Other"),
    (15, "misc", "Misc.(visa, insurance, others)", "HKD"),
    (16, "misc", "Misc.(visa, insurance, others)", "RMB"),
    (17, "misc", "Misc.(visa, insurance, others)", "Other"),
]
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def relative_to_folder(path: Path | None, folder: Path) -> str | None:
    if not path:
        return None
    try:
        return os.path.relpath(path, folder)
    except ValueError:
        return str(path)


def load_workbook_from_onedrive_safe(path: Path) -> Any:
    """Copy first so openpyxl can read OneDrive reparse/placeholder files."""
    with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        shutil.copyfile(path, temp_path)
        return openpyxl.load_workbook(temp_path, data_only=False)
    finally:
        temp_path.unlink(missing_ok=True)


def find_travel_workbook(folder: Path, explicit: Path | None = None) -> Path:
    if explicit:
        if explicit.exists():
            return explicit
        raise FileNotFoundError(f"Travel workbook not found: {explicit}")

    preferred = sorted(folder.glob("差旅報銷清單_行程資料列表Reimbursement for travel expenses*.xlsx"))
    if preferred:
        return preferred[0]

    fallback = sorted(folder.glob("*travel expenses*.xlsx"))
    if fallback:
        return fallback[0]

    raise FileNotFoundError(f"No travel reimbursement workbook found in {folder}")


def parse_date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = cell_text(value)
    if not text:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def serializable_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def parse_amount(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = cell_text(value)
    if not text or text.startswith("="):
        return None
    text = re.sub(r"(HKD|RMB|CNY|USD|JPY|MOP|NTD)", "", text, flags=re.IGNORECASE)
    text = text.replace(",", "").replace("￥", "").replace("¥", "").replace("$", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def value_after_label(value: Any, *labels: str) -> str:
    text = cell_text(value)
    for label in labels:
        match = re.match(rf"^{re.escape(label)}\s*:?\s*(.*)$", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return text


def find_total_row(worksheet: Any) -> int:
    for row in range(1, worksheet.max_row + 1):
        if cell_text(worksheet.cell(row, 1).value).startswith("Total:"):
            return row
    return worksheet.max_row + 1


def parse_profile(worksheet: Any) -> dict[str, str]:
    return {
        "name": value_after_label(worksheet["A3"].value, "Applicant"),
        "bank": value_after_label(worksheet["I3"].value, "Bank"),
        "account": value_after_label(worksheet["I5"].value, "Bank Account Number"),
        "purpose": value_after_label(worksheet["A6"].value, "Purpose of the trip"),
    }


def parse_expenses(worksheet: Any) -> list[dict[str, Any]]:
    total_row = find_total_row(worksheet)
    expenses: list[dict[str, Any]] = []
    for row_number in range(10, total_row):
        expense_date = parse_date_value(worksheet.cell(row_number, 1).value)
        destination = cell_text(worksheet.cell(row_number, 2).value)
        if not expense_date and not destination:
            continue
        raw_row = {
            get_column_letter(column): serializable_cell(worksheet.cell(row_number, column).value)
            for column in range(1, 18)
        }
        for column, category, category_label, currency in EXPENSE_COLUMNS:
            amount = parse_amount(worksheet.cell(row_number, column).value)
            if amount is None:
                continue
            expenses.append(
                {
                    "source_row_index": row_number,
                    "expense_date": expense_date,
                    "destination": destination,
                    "category": category,
                    "category_label": category_label,
                    "currency": currency,
                    "amount": amount,
                    "raw": {
                        "worksheet": worksheet.title,
                        "row": row_number,
                        "cell": f"{get_column_letter(column)}{row_number}",
                        "row_values": raw_row,
                    },
                }
            )
    return expenses


def parse_itinerary(worksheet: Any) -> list[dict[str, Any]]:
    itineraries: list[dict[str, Any]] = []
    for row_number in range(2, worksheet.max_row + 1):
        index_value = worksheet.cell(row_number, 1).value
        trip_date = parse_date_value(worksheet.cell(row_number, 2).value)
        origin = cell_text(worksheet.cell(row_number, 3).value)
        destination = cell_text(worksheet.cell(row_number, 4).value)
        purpose = cell_text(worksheet.cell(row_number, 5).value)
        if not any([trip_date, origin, destination, purpose]):
            continue
        try:
            itinerary_index = int(index_value)
        except (TypeError, ValueError):
            itinerary_index = len(itineraries) + 1
        itineraries.append(
            {
                "itinerary_index": itinerary_index,
                "trip_date": trip_date,
                "origin": origin,
                "destination": destination,
                "purpose": purpose,
                "raw": {
                    "worksheet": worksheet.title,
                    "row": row_number,
                    "index": serializable_cell(index_value),
                    "date": serializable_cell(worksheet.cell(row_number, 2).value),
                },
            }
        )
    return itineraries


def image_dimensions(path: Path) -> tuple[int | None, int | None]:
    try:
        with Image.open(path) as image:
            return image.size
    except (OSError, UnidentifiedImageError):
        return None, None


def file_metadata(path: Path | None, batch_folder: Path) -> dict[str, Any]:
    if not path or not path.exists() or not path.is_file():
        return {
            "actual_path": str(path) if path else None,
            "relative_path": relative_to_folder(path, batch_folder) if path else None,
            "file_name": path.name if path else None,
            "file_size": None,
            "sha256": None,
            "width": None,
            "height": None,
        }
    width, height = image_dimensions(path) if path.suffix.lower() in IMAGE_SUFFIXES else (None, None)
    return {
        "actual_path": str(path),
        "relative_path": relative_to_folder(path, batch_folder),
        "file_name": path.name,
        "file_size": path.stat().st_size,
        "sha256": state_db.sha256_file(path),
        "width": width,
        "height": height,
    }


def docx_media_details(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with zipfile.ZipFile(path) as archive:
        media = [name for name in archive.namelist() if name.startswith("word/media/")]
    return {
        "embedded_media_count": len(media),
        "embedded_media": media,
    }


def collect_travel_evidence(batch_folder: Path) -> list[dict[str, Any]]:
    evidence: list[dict[str, Any]] = []
    docx_path = batch_folder / "差旅.docx"
    if docx_path.exists():
        evidence.append(
            {
                "evidence_index": 0,
                "evidence_kind": "travel_evidence_docx_bundle",
                "expected_filename": docx_path.name,
                **file_metadata(docx_path, batch_folder),
                "raw_path": None,
                "source_path": relative_to_folder(docx_path.parent, batch_folder),
                "capture_method": "manual_docx_image_bundle",
                "validation_status": "valid",
                "warnings": [],
                "details": docx_media_details(docx_path),
            }
        )

    travel_folder = batch_folder / "差旅"
    images = sorted(path for path in travel_folder.glob("*") if path.suffix.lower() in IMAGE_SUFFIXES)
    for index, path in enumerate(images, 1):
        width, height = image_dimensions(path)
        warnings = [] if width and height else ["image file could not be opened"]
        evidence.append(
            {
                "evidence_index": index,
                "evidence_kind": "travel_evidence_image",
                "expected_filename": path.name,
                **file_metadata(path, batch_folder),
                "raw_path": None,
                "source_path": relative_to_folder(travel_folder, batch_folder),
                "capture_method": "manual_travel_payment_or_route_screenshot",
                "validation_status": "valid" if not warnings else "warning",
                "warnings": warnings,
                "details": {
                    "folder": str(travel_folder),
                },
            }
        )

    if not evidence:
        evidence.append(
            {
                "evidence_index": 1,
                "evidence_kind": "travel_evidence_image",
                "expected_filename": None,
                "actual_path": None,
                "relative_path": None,
                "file_name": None,
                "file_size": None,
                "sha256": None,
                "width": None,
                "height": None,
                "raw_path": None,
                "source_path": relative_to_folder(travel_folder, batch_folder),
                "capture_method": "manual_travel_payment_or_route_screenshot",
                "validation_status": "missing",
                "warnings": ["No travel evidence folder or docx image bundle found"],
                "details": {
                    "folder": str(travel_folder),
                    "docx": str(docx_path),
                },
            }
        )
    return evidence


def amount_totals(expenses: list[dict[str, Any]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for expense in expenses:
        currency = str(expense["currency"])
        totals[currency] = round(totals.get(currency, 0.0) + float(expense["amount"]), 2)
    return totals


def build_manifest(
    *,
    workbook_path: Path,
    profile: dict[str, str],
    expenses: list[dict[str, Any]],
    itinerary: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
) -> dict[str, Any]:
    valid_evidence = [row for row in evidence if row.get("validation_status") == "valid"]
    return {
        "schema": "prepare-reimbursements.travel.v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": str(workbook_path),
        "profile": profile,
        "summary": {
            "expense_records": len(expenses),
            "expense_source_rows": len({expense["source_row_index"] for expense in expenses}),
            "amount_totals": amount_totals(expenses),
            "itinerary_rows": len(itinerary),
            "evidence_records": len(evidence),
            "valid_evidence_records": len(valid_evidence),
        },
        "expenses": expenses,
        "itinerary": itinerary,
        "evidence": [
            {
                "index": row["evidence_index"],
                "kind": row["evidence_kind"],
                "relative_path": row.get("relative_path"),
                "status": row.get("validation_status"),
                "size": [row.get("width"), row.get("height")] if row.get("width") and row.get("height") else None,
                "warnings": row.get("warnings") or [],
            }
            for row in evidence
        ],
    }


def artifact_for(path: Path, batch_folder: Path, kind: str, details: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "artifact_kind": kind,
        "path": str(path),
        "relative_path": relative_to_folder(path, batch_folder),
        "file_size": path.stat().st_size if path.exists() and path.is_file() else None,
        "sha256": state_db.sha256_file(path) if path.exists() and path.is_file() else None,
        "details": details or {},
    }


def sync_batch(
    *,
    batch_folder: Path,
    workbook_path: Path,
    db_path: Path,
    out_dir: Path,
    snapshot_path: Path | None,
    overrides: dict[str, str],
) -> dict[str, Any]:
    workbook = load_workbook_from_onedrive_safe(workbook_path)
    expense_sheet = workbook["差旅報銷清單"]
    itinerary_sheet = workbook["行程資料列表"]

    profile = parse_profile(expense_sheet)
    profile.update({key: value for key, value in overrides.items() if value})
    expenses = parse_expenses(expense_sheet)
    itinerary = parse_itinerary(itinerary_sheet)
    evidence = collect_travel_evidence(batch_folder)

    out_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = out_dir / MANIFEST_NAME
    manifest = build_manifest(
        workbook_path=workbook_path,
        profile=profile,
        expenses=expenses,
        itinerary=itinerary,
        evidence=evidence,
    )
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    with state_db.connect(db_path) as connection:
        existing = connection.execute(
            "SELECT * FROM batches WHERE batch_folder = ?",
            (str(batch_folder),),
        ).fetchone()
        reimbursement_type = state_db.merge_reimbursement_type(
            existing["reimbursement_type"] if existing else None,
            "travel",
        )
        merged_profile = state_db.json_loads(existing["profile_json"], {}) if existing else {}
        merged_profile.update(profile)
        travel_summary = {
            **manifest["summary"],
            "source_workbook": str(workbook_path),
            "manifest": str(manifest_path),
        }
        if existing and existing["reimbursement_type"] != "travel":
            merged_summary = state_db.json_loads(existing["summary_json"], {})
            merged_summary["travel"] = travel_summary
            source_manifest_path = Path(existing["source_manifest_path"]) if existing["source_manifest_path"] else manifest_path
            source_export_path = Path(existing["source_export_path"]) if existing["source_export_path"] else workbook_path
        else:
            merged_summary = {**manifest["summary"], "travel": travel_summary}
            source_manifest_path = manifest_path
            source_export_path = workbook_path
        batch_id = state_db.upsert_batch(
            connection,
            batch_folder=batch_folder,
            reimbursement_type=reimbursement_type,
            source_manifest_path=source_manifest_path,
            source_export_path=source_export_path,
            profile=merged_profile,
            summary=merged_summary,
        )
        expense_count = state_db.replace_travel_expenses(connection, batch_id=batch_id, expenses=expenses)
        itinerary_count = state_db.replace_travel_itineraries(connection, batch_id=batch_id, itineraries=itinerary)
        evidence_count = state_db.replace_travel_evidence_files(
            connection,
            batch_id=batch_id,
            evidence_files=evidence,
        )
        warnings = [
            warning
            for row in evidence
            for warning in row.get("warnings", [])
        ]
        state_db.replace_validation_results(
            connection,
            batch_id=batch_id,
            tool="scripts/sync_travel_reimbursement_state.py",
            results=[
                {
                    "scope": "travel_evidence",
                    "status": "valid" if not warnings else "warning",
                    "warnings": warnings,
                    "details": {
                        "evidence_records": evidence_count,
                        "valid_evidence_records": manifest["summary"]["valid_evidence_records"],
                    },
                }
            ],
        )
        artifact_count = state_db.upsert_artifacts(
            connection,
            batch_id=batch_id,
            artifacts=[
                artifact_for(manifest_path, batch_folder, "travel_manifest"),
            ],
        )

        snapshot_written = None
        if snapshot_path:
            payload = state_db.snapshot(connection, batch_id=batch_id, db_path=db_path)
            snapshot_path.parent.mkdir(parents=True, exist_ok=True)
            snapshot_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            snapshot_written = str(snapshot_path)

        connection.commit()

    return {
        "database": str(db_path),
        "snapshot": snapshot_written,
        "batch": str(batch_folder),
        "source_workbook": str(workbook_path),
        "profile": merged_profile,
        "expenses": expense_count,
        "expense_source_rows": manifest["summary"]["expense_source_rows"],
        "itinerary_rows": itinerary_count,
        "evidence_records": evidence_count,
        "artifacts": artifact_count,
        "amount_totals": manifest["summary"]["amount_totals"],
        "warnings": warnings,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--folder", type=Path, required=True, help="Reimbursement batch folder")
    parser.add_argument("--workbook", type=Path, help="Travel workbook. Defaults to the workbook in the batch folder")
    parser.add_argument("--db", type=Path, help="SQLite state database. Defaults to <folder>/generated/reimbursement-state.sqlite3")
    parser.add_argument("--out-dir", type=Path, help="Output directory. Defaults to <folder>/generated")
    parser.add_argument(
        "--snapshot",
        type=Path,
        help="Review-friendly JSON snapshot. Defaults to <folder>/generated/reimbursement-state.snapshot.json",
    )
    parser.add_argument("--no-snapshot", action="store_true", help="Do not write the JSON snapshot")
    parser.add_argument("--name")
    parser.add_argument("--bank")
    parser.add_argument("--account")
    parser.add_argument("--purpose")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    batch_folder = args.folder.resolve()
    out_dir = (args.out_dir or batch_folder / "generated").resolve()
    db_path = (args.db or out_dir / DEFAULT_DB_NAME).resolve()
    workbook_path = find_travel_workbook(batch_folder, args.workbook).resolve()
    snapshot_path = None if args.no_snapshot else (args.snapshot or out_dir / DEFAULT_SNAPSHOT_NAME).resolve()
    overrides = {
        "name": args.name,
        "bank": args.bank,
        "account": args.account,
        "purpose": args.purpose,
    }

    result = sync_batch(
        batch_folder=batch_folder,
        workbook_path=workbook_path,
        db_path=db_path,
        out_dir=out_dir,
        snapshot_path=snapshot_path,
        overrides=overrides,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
